import os
import re
import camelot
# prevent Camelot's aggressive atexit removal which on Windows can produce PermissionError spam
# we will manage temp cleanup ourselves
try:
    camelot.utils.remove_tempdir = lambda *args, **kwargs: None
except Exception:
    pass

import pdfplumber
from pdf2image import convert_from_path
import pytesseract
import cv2
import numpy as np
import pandas as pd
from datetime import datetime
from dateutil import parser as dateparser
from concurrent.futures import ProcessPoolExecutor, as_completed
import tempfile
import shutil
import time
import gc

PENALTY_RATE_PER_DAY = 0.0005
OUTPUT_FILE = "muster_extracted_minimal.xlsx"

MAX_WORKERS = max(2, os.cpu_count() - 1)
OCR_DPI_PRIMARY = 220
OCR_DPI_FALLBACK = 300
TESS_CONFIG = "--oem 1 --psm 6"

# ---------------- UTILITY ----------------

def safe_date_parse(s):
    if not s:
        return None
    try:
        return dateparser.parse(s, dayfirst=True).date()
    except:
        return None

def extract_header_info_from_text(text):
    work_code = None
    date_to = None

    wc = re.search(r"Work Code\s*[:\-]?\s*([A-Za-z0-9\/\-]+)", text, flags=re.I)
    dt = re.search(r"Date To\s*[:\-]?\s*(\d{1,2}/\d{1,2}/\d{4})", text, flags=re.I)

    if wc:
        work_code = wc.group(1).strip()
    if dt:
        date_to = dt.group(1).strip()

    return work_code, date_to

def normalize_amount(val):
    if val is None:
        return None
    s = str(val)
    s = s.replace(",", "").replace("₹", "").replace("Rs.", "").replace("Rs", "")
    m = re.search(r"(\d+(?:\.\d+)?)", s)
    return float(m.group(1)) if m else None

def compute_delay_penalty(closing_date, credited_date, amount):
    closing = safe_date_parse(closing_date)
    credited = safe_date_parse(credited_date)
    if closing is None or credited is None:
        return None, None
    diff = (credited - closing).days
    delay = max(0, diff - 15)
    penalty = round((amount or 0) * PENALTY_RATE_PER_DAY * delay, 2)
    return delay, penalty

# ------------------- Helpers for temp cleanup -------------------

def _remove_dir_with_retries(path, attempts=6, pause=0.25):
    """Try to remove `path` with a few retries to handle transient file locks on Windows."""
    if not path or not os.path.exists(path):
        return
    for i in range(attempts):
        try:
            shutil.rmtree(path)
            return
        except PermissionError:
            # give time for file handles to be released
            gc.collect()
            time.sleep(pause)
        except Exception:
            # last resort: ignore
            try:
                gc.collect()
                time.sleep(pause)
                shutil.rmtree(path)
                return
            except Exception:
                return

# --------- CAMeLOT EXTRACTION (safe cleanup) ----------

def extract_with_camelot(pdf_path):
    """
    Try lattice then stream. Safely release Camelot table objects and collect DataFrame.
    """
    all_dfs = []
    for flavor in ("lattice", "stream"):
        tables = None
        try:
            tables = camelot.read_pdf(pdf_path, pages="all", flavor=flavor, strip_text="\n")
        except Exception:
            tables = None

        if not tables or getattr(tables, "n", 0) == 0:
            # no tables found for this flavor
            continue

        # pick best table by r*c as before
        best = None
        best_score = -1
        for t in tables:
            try:
                r, c = t.shape
            except Exception:
                continue
            if r * c > best_score:
                best = t
                best_score = r * c

        if best is None:
            # still nothing usable
            # ensure we delete table objects to free handles
            for t in tables:
                try:
                    del t
                except Exception:
                    pass
            gc.collect()
            continue

        df = best.df.copy()
        # free the table object(s)
        try:
            for t in tables:
                del t
        except Exception:
            pass
        gc.collect()

        # normalize empty rows
        df = df.replace(r"^\s*$", np.nan, regex=True)

        # convert first row to header if more than 1 row
        if df.shape[0] > 1:
            df.columns = df.iloc[0].fillna("").astype(str)
            df = df[1:].reset_index(drop=True)
        else:
            df.columns = [f"col_{i}" for i in range(df.shape[1])]

        # remove DAILY ATTENDANCE column (case-insensitive)
        drop_cols = [c for c in df.columns if c and "daily" in str(c).lower()]
        if drop_cols:
            df = df.drop(columns=drop_cols, errors="ignore")

        return df

    return None

# ---------------- OCR (safe pdf2image temp handling) ----------------

def ocr_page_images(pdf_path, dpi):
    """
    Use a dedicated temp dir for convert_from_path, return list of PIL images.
    We'll return PIL images and the tempdir path so caller can close images and delete folder.
    """
    tempdir = tempfile.mkdtemp(prefix="muster_pdf2img_")
    try:
        images = convert_from_path(pdf_path, dpi=dpi, fmt="png", output_folder=tempdir)
        # convert_from_path sometimes returns generator-like list; ensure it's a real list
        images = list(images)
        return images, tempdir
    except Exception:
        # cleanup if convert fails
        _remove_dir_with_retries(tempdir)
        raise

def extract_table_via_ocr(pdf_path, dpi=OCR_DPI_PRIMARY):
    ocr_rows = []
    images = []
    tempdir = None
    try:
        images, tempdir = ocr_page_images(pdf_path, dpi=dpi)
    except Exception as e:
        # fallback: return empty list if convert fails
        return []

    try:
        for pil_img in images:
            try:
                img = cv2.cvtColor(np.array(pil_img), cv2.COLOR_RGB2BGR)
            except Exception:
                # skip problematic page
                continue
            finally:
                # close PIL image to release any file handles
                try:
                    pil_img.close()
                except Exception:
                    pass

            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            _, th = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
            th = 255 - th

            # detect lines (horizontal & vertical) to find table structure
            horizontal = th.copy()
            h_size = max(10, horizontal.shape[1] // 30)
            h_str = cv2.getStructuringElement(cv2.MORPH_RECT, (h_size, 1))
            horizontal = cv2.erode(horizontal, h_str)
            horizontal = cv2.dilate(horizontal, h_str)

            vertical = th.copy()
            v_size = max(10, vertical.shape[0] // 30)
            v_str = cv2.getStructuringElement(cv2.MORPH_RECT, (1, v_size))
            vertical = cv2.erode(vertical, v_str)
            vertical = cv2.dilate(vertical, v_str)

            mask = cv2.add(horizontal, vertical)

            contours, _ = cv2.findContours(mask, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
            boxes = []
            for cnt in contours:
                x, y, w, h = cv2.boundingRect(cnt)
                if w > 40 and h > 10:
                    boxes.append((x, y, w, h))

            # fallback: full-page OCR if no table boxes detected
            if not boxes:
                text = pytesseract.image_to_string(img, config=TESS_CONFIG)
                lines = [ln.strip() for ln in text.split("\n") if ln.strip()]
                # filter out daily attendance lines
                lines = [ln for ln in lines if "daily" not in ln.lower()]
                ocr_rows.extend(lines)
                continue

            # group boxes into rows by y coordinate
            boxes = sorted(boxes, key=lambda b: (b[1], b[0]))
            rows = []
            current = []
            last_y = -9999

            for b in boxes:
                x, y, w, h = b
                if abs(y - last_y) <= 10:
                    current.append(b)
                else:
                    if current:
                        rows.append(current)
                    current = [b]
                last_y = y
            if current:
                rows.append(current)

            for row in rows:
                row = sorted(row, key=lambda b: b[0])
                values = []
                for (x, y, w, h) in row:
                    crop = img[y:y+h, x:x+w]
                    text = pytesseract.image_to_string(crop, config=TESS_CONFIG).strip()
                    values.append(text)
                # join with delimiter and filter lines that look like daily attendance
                joined = " ||| ".join(values)
                if "daily" in joined.lower():
                    continue
                ocr_rows.append(joined)
    finally:
        # ensure tempdir is removed (with retries) and free memory
        gc.collect()
        _remove_dir_with_retries(tempdir)

    return ocr_rows

# ---------------- PARSING ----------------

def robust_sno_from_cells(row_dict, header_map):
    for h in header_map:
        if re.search(r"(s\.?no|serial|sl\.?no)", str(h), re.I):
            val = str(row_dict.get(h, "")).strip()
            m = re.match(r"^\D*(\d+)\D*$", val)
            if m:
                return m.group(1)

    for h in header_map:
        val = str(row_dict.get(h, "")).strip()
        if val.isdigit():
            return val
    return None

def extract_name_like(row_dict, header_map, all_text):
    for h in header_map:
        if re.search(r"name|reg.?no", str(h), re.I):
            name = str(row_dict.get(h, "")).strip()
            if name:
                return name

    tmp = re.sub(r"^\s*\d+\s*", "", all_text).strip()
    tmp = re.split(r"\s{2,}|\|\|\|", tmp)[0]
    return tmp[:200].strip()

def extract_amount_like(row_dict, header_map, all_text):
    for h in header_map:
        if "amount" in str(h).lower():
            amt = normalize_amount(row_dict.get(h))
            if amt is not None:
                return amt
    nums = re.findall(r"\b\d{3,7}\b", all_text.replace(",", ""))
    return float(max(nums, key=lambda x: int(x))) if nums else None

def extract_credited_date(row_dict, header_map, all_text):
    for h in header_map:
        if re.search(r"(credited|a/c).*date", str(h), re.I):
            dates = re.findall(r"\d{1,2}/\d{1,2}/\d{4}", str(row_dict.get(h, "")))
            if dates:
                return dates[-1]
    dates = re.findall(r"\d{1,2}/\d{1,2}/\d{4}", all_text)
    return dates[-1] if dates else None

def parse_row_from_table(row_dict, header_map, work_code, date_to):
    all_text = " ".join(str(row_dict.get(h, "")) for h in header_map)

    return {
        "S.No": robust_sno_from_cells(row_dict, header_map),
        "Name/RegNo": extract_name_like(row_dict, header_map, all_text),
        "Amount Due": extract_amount_like(row_dict, header_map, all_text),
        "Work Code": work_code,
        "Work Closing Date": date_to,
        "A/c Credited Date": extract_credited_date(row_dict, header_map, all_text)
    }

def parse_row_from_ocr(line, work_code, date_to):
    line = line.replace("daily", "").strip()

    m = re.match(r"^\s*(\d+)", line)
    sno = m.group(1) if m else None

    dates = re.findall(r"\d{1,2}/\d{1,2}/\d{4}", line)
    credited = dates[-1] if dates else None

    nums = re.findall(r"\b\d{3,7}\b", line.replace(",", ""))
    amount = float(max(nums, key=lambda x: int(x))) if nums else None

    parts = line.split("|||")
    name = parts[1].strip() if len(parts) > 1 else line

    return {
        "S.No": sno,
        "Name/RegNo": name,
        "Amount Due": amount,
        "Work Code": work_code,
        "Work Closing Date": date_to,
        "A/c Credited Date": credited
    }

# ---------------- PDF PROCESS ----------------

def process_pdf(pdf_path):
    results = []
    full_text = ""

    try:
        with pdfplumber.open(pdf_path) as pdf:
            for p in pdf.pages:
                t = p.extract_text()
                if t:
                    full_text += t + "\n"
    except Exception:
        pass

    work_code, date_to = extract_header_info_from_text(full_text)

    df = extract_with_camelot(pdf_path)
    if df is not None and df.shape[0] > 0:
        header_map = list(df.columns)
        for _, row in df.iterrows():
            parsed = parse_row_from_table(row.to_dict(), header_map, work_code, date_to)
            results.append(parsed)
        return results

    ocr_lines = extract_table_via_ocr(pdf_path, dpi=OCR_DPI_PRIMARY)
    if len(ocr_lines) < 3:
        ocr_lines = extract_table_via_ocr(pdf_path, dpi=OCR_DPI_FALLBACK)

    for ln in ocr_lines:
        results.append(parse_row_from_ocr(ln, work_code, date_to))

    return results

# ---------------- FOLDER PROCESS ----------------

def _postprocess_rows(rows, src_file):
    post = []
    seq = 1
    for r in rows:
        r["Source_File"] = src_file

        if not r.get("S.No") or not str(r["S.No"]).isdigit():
            r["S.No"] = str(seq)
        seq += 1

        delay, penalty = compute_delay_penalty(
            r.get("Work Closing Date"),
            r.get("A/c Credited Date"),
            r.get("Amount Due")
        )
        r["Delay (after 15 days)"] = delay
        r["Penalty 0.05% per day of delay"] = penalty

        if delay and delay > 0:
            post.append(r)

    return post

def _safe_process_file(folder, f):
    path = os.path.join(folder, f)
    try:
        rows = process_pdf(path)
        return _postprocess_rows(rows, f)
    except Exception as e:
        print("Error reading", f, ":", e)
        return []

def process_folder(folder):
    pdfs = [f for f in os.listdir(folder) if f.lower().endswith(".pdf")]
    pdfs.sort()

    all_rows = []

    with ProcessPoolExecutor(max_workers=MAX_WORKERS) as ex:
        futures = {ex.submit(_safe_process_file, folder, f): f for f in pdfs}

        for fut in as_completed(futures):
            f = futures[fut]
            try:
                chunk = fut.result()
                all_rows.extend(chunk)
                print(f"Processed: {f} | +{len(chunk)} delayed rows")
            except Exception as e:
                print(f"Failed: {f} ({e})")

    if not all_rows:
        cols = [
            "S.No",
            "Name/RegNo",
            "Amount Due",
            "Work Code",
            "Work Closing Date",
            "A/c Credited Date",
            "Delay (after 15 days)",
            "Penalty 0.05% per day of delay",
            "Source_File"
        ]
        pd.DataFrame(columns=cols).to_excel(OUTPUT_FILE, index=False)
        print("No delayed rows found.")
        return

    df = pd.DataFrame(all_rows)

    required_cols = [
        "S.No",
        "Name/RegNo",
        "Amount Due",
        "Work Code",
        "Work Closing Date",
        "A/c Credited Date",
        "Delay (after 15 days)",
        "Penalty 0.05% per day of delay",
        "Source_File"
    ]

    df = df[[c for c in required_cols if c in df.columns]]
    df.to_excel(OUTPUT_FILE, index=False)

    # ----- Excel Formatting -----
    from openpyxl import load_workbook
    from openpyxl.styles import Border, Side, Alignment, Font, PatternFill

    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active

    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # header styling
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4F81BD")
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # data formatting
    for column_cells in ws.columns:
        max_length = 0
        col_letter = column_cells[0].column_letter

        for cell in column_cells:
            cell.border = border
            try:
                max_length = max(max_length, len(str(cell.value)) if cell.value else 0)
            except:
                pass

        ws.column_dimensions[col_letter].width = min(80, max(12, max_length + 4))

    wb.save(OUTPUT_FILE)
    print("DONE — Saved:", OUTPUT_FILE)

# ---------------- MAIN ----------------

if __name__ == "__main__":
    folder = input("Enter folder path containing Muster Roll PDF files:\n> ").strip()
    if os.path.isdir(folder):
        process_folder(folder)
    else:
        print("Invalid folder path.")
